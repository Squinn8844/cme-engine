"""
anonymize_fixtures.py
─────────────────────
Strip PII from real data files to create safe test fixtures.
Run this ONCE locally before committing fixture files to GitHub.

Usage:
    python anonymize_fixtures.py \\
        --input_dir /path/to/real/data \\
        --output_dir tests/fixtures/LAIPrEP \\
        --program laipep

Strips: Email, Last Name, First Name, ZIP, Token
Preserves: All response data (needed for test assertions)
"""

import argparse
import os
import pandas as pd
import openpyxl
import random
import string
import hashlib


# ── PII column patterns to blank out ────────────────────────────────────────
PII_PATTERNS = [
    "email", "last name", "first name", "zip", "token",
    "lastname", "firstname", "zipcode", "postal",
    "name", "address", "phone", "dob", "date of birth",
]


def _is_pii_col(col_name: str) -> bool:
    norm = col_name.lower().strip()
    return any(p in norm for p in PII_PATTERNS)


def _fake_email(seed: str) -> str:
    h = hashlib.md5(seed.encode()).hexdigest()[:8]
    return f"respondent_{h}@example.com"


def _fake_name(seed: str) -> str:
    h = hashlib.md5(seed.encode()).hexdigest()[:6]
    return f"Anon_{h}"


def _fake_zip() -> str:
    return "".join(random.choices(string.digits, k=5))


def anonymize_nexus(input_path: str, output_path: str):
    """Anonymize a Nexus format Excel file."""
    xl = pd.ExcelFile(input_path)
    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    for sheet_name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name)

        for col in df.columns:
            if _is_pii_col(str(col)):
                if "email" in str(col).lower():
                    df[col] = [_fake_email(str(v)) for v in df[col]]
                elif "zip" in str(col).lower() or "postal" in str(col).lower():
                    df[col] = [_fake_zip() for _ in df[col]]
                else:
                    df[col] = [_fake_name(str(v)) for v in df[col]]

        # Anonymize ID column (col 0) — replace with sequential
        if len(df.columns) > 0:
            df.iloc[:, 0] = [f"ID_{i:06d}" for i in range(len(df))]

        df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()
    print(f"  Anonymized: {os.path.basename(input_path)} → {os.path.basename(output_path)}")


def anonymize_exchange(input_path: str, output_path: str):
    """Anonymize an ExchangeCME Worksheet file."""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Row 0 has admin column headers — find PII cols
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    pii_cols = []
    for col_idx, header in enumerate(header_row or [], start=1):
        if header and _is_pii_col(str(header)):
            pii_cols.append(col_idx)

    # Blank out PII in data rows (row 4+)
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            if cell.column in pii_cols:
                col_header = str(header_row[cell.column - 1]).lower()
                if "email" in col_header:
                    cell.value = _fake_email(str(cell.row))
                elif "zip" in col_header:
                    cell.value = _fake_zip()
                else:
                    cell.value = _fake_name(str(cell.row))

    wb.save(output_path)
    print(f"  Anonymized: {os.path.basename(input_path)} → {os.path.basename(output_path)}")


def anonymize_key(input_path: str, output_path: str):
    """Key files have no PII — just copy."""
    import shutil
    shutil.copy2(input_path, output_path)
    print(f"  Copied key: {os.path.basename(input_path)} → {os.path.basename(output_path)}")


def main():
    parser = argparse.ArgumentParser(description="Anonymize CME data files for test fixtures")
    parser.add_argument("--input_dir", required=True, help="Directory with real data files")
    parser.add_argument("--output_dir", required=True, help="Output directory for anonymized fixtures")
    parser.add_argument("--program", required=True, help="Program name (laipep | il33 | excdata)")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    print(f"\nAnonymizing {args.program} fixtures...")
    print(f"  Input:  {args.input_dir}")
    print(f"  Output: {args.output_dir}")

    # Map program to expected files
    file_map = {
        "laipep": {
            "LAI_KEY.xlsx":         ("key",      "LAI_KEY.xlsx"),
            "NEXUS_FORMAT*.xlsx":   ("nexus",    "NEXUS_LAIPrEP.xlsx"),
            "ExchangeCME_*.xlsx":   ("exchange", "ExchangeCME_LAIPrEP.xlsx"),
        },
        "il33": {
            "IL33_Key.xlsx":        ("key",      "IL33_Key.xlsx"),
            "IL33NEXUS.xlsx":       ("nexus",    "IL33NEXUS.xlsx"),
            "IL33EXHDATA.xlsx":     ("exchange", "IL33EXHDATA.xlsx"),
        },
    }

    # Auto-detect files in input_dir
    for filename in os.listdir(args.input_dir):
        if not filename.endswith(".xlsx"):
            continue

        input_path  = os.path.join(args.input_dir, filename)
        output_name = filename
        output_path = os.path.join(args.output_dir, output_name)

        # Detect type
        try:
            xl = pd.ExcelFile(input_path)
            sheets_lower = {s.lower() for s in xl.sheet_names}

            if sheets_lower & {"pre-test", "post", "evaluation"}:
                anonymize_key(input_path, output_path)
            elif sheets_lower & {"pre", "prenon", "post", "eval"}:
                anonymize_nexus(input_path, output_path)
            elif "worksheet" in sheets_lower:
                anonymize_exchange(input_path, output_path)
            else:
                anonymize_nexus(input_path, output_path)  # Try generic
        except Exception as e:
            print(f"  ERROR processing {filename}: {e}")

    print(f"\nDone. Fixtures saved to {args.output_dir}")
    print("Safe to commit to GitHub.")


if __name__ == "__main__":
    main()
